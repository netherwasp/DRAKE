# DRAKE 3.0
import sys
from datetime import date
from datetime import datetime
import drakeDBinit as drakeDB
import sqlite3
import calendar
import os
from openpyxl import Workbook, load_workbook

from PyQt5 import QtGui
from PyQt5.QtWidgets import QComboBox, QFrame, QWidget, QMainWindow, QPushButton, QLineEdit, QLabel, QSpacerItem, \
    QHBoxLayout, QVBoxLayout, QGridLayout, QTableView, QSizePolicy, QApplication, QHeaderView, QDateEdit, QMessageBox, \
    QSpinBox, QAbstractItemView, QAction
from PyQt5.QtCore import Qt, QAbstractItemModel, QRegExp, QAbstractTableModel, QDate, pyqtSlot
from PyQt5.QtGui import QFont, QStandardItemModel, QStandardItem, QIcon, QFont
from qt_material import apply_stylesheet




class drakeMain(QMainWindow):
    def __init__(self):
        drakeDB.drakeDbfetch()
        curr_date = datetime.now()
        super(drakeMain, self).__init__()
        self.drakeMainWidget = QWidget()
        self.setCentralWidget(self.drakeMainWidget)

        self.setWindowIcon(QIcon("resources/bearded-dragon.png"))
        self.setWindowTitle("D.R.A.K.E.")

        self.setMinimumSize(1280, 640)

        self.drakeGridLayout = QGridLayout(self.drakeMainWidget)

        self.genFont = QFont()
        self.genFont.setBold(True)
        self.genFont.setWeight(100)

        # ENTRY FRAME
        self.entryFrame = QFrame()
        self.entryFrame.setContentsMargins(0, 0, 0, 0)
        self.entryFrame.setMaximumHeight(80)

        self.entryGridLayout = QGridLayout(self.entryFrame)
        self.entryGridLayout.setContentsMargins(5, 5, 5, 5)

        self.dateLabel = QLabel()
        self.dateLabel.setText(" DATE: ")
        self.dateLabel.setFixedWidth(40)

        year = int(curr_date.strftime("%Y"))
        self.year = QComboBox()
        self.year.addItems([str(year) for year in range(2023, year + 50)])
        self.year.setMinimumHeight(40)
        self.year.setMinimumWidth(80)
        self.year.setCurrentText(str(year))
        self.year.setFont(self.genFont)

        self.month = QComboBox()
        self.month.addItems([str(month) for month in range(1, 13)])
        self.month.setMinimumHeight(40)
        self.month.setMinimumWidth(60)
        self.month.setCurrentText(curr_date.strftime("%#m"))
        self.month.setFont(self.genFont)

        self.day = QComboBox()
        self.day.setMinimumHeight(40)
        self.day.setMinimumWidth(60)
        self.setDayItemsMethod()
        self.day.setCurrentText(curr_date.strftime("%#d"))
        self.day.setFont(self.genFont)

        self.date = f"{self.month.currentText()}/{self.day.currentText()}/{self.year.currentText()}"

        self.entryField = QLineEdit()
        self.entryField.setPlaceholderText("ENTER ID NUMBER HERE...")
        self.entryField.setMinimumHeight(40)
        self.entryField.setMinimumWidth(280)
        self.entryField.setMaxLength(10)
        self.entryField.setValidator(QtGui.QRegExpValidator(QRegExp(r"^[0-9]*$")))
        self.entryField.setFont(self.genFont)

        self.entryButton = QPushButton()
        self.entryButton.setMinimumHeight(40)
        self.entryButton.setFixedWidth(80)
        self.entryButton.setText("LOG")
        self.entryButton.setFont(self.genFont)

        self.entryButtonAction = QAction()
        self.entryButtonAction.setShortcuts([Qt.Key_Enter,Qt.Key_Return])
        self.entryButtonAction.triggered.connect(self.entryButtonMethod)
        self.entryButton.addAction(self.entryButtonAction)

        self.entryComboBox = QComboBox()
        entryComboItems = ["TIME-IN", "TIME-OUT"]
        self.entryComboBox.addItems(entryComboItems)
        self.entryComboBox.setFixedWidth(100)
        self.entryComboBox.setMinimumHeight(40)
        self.entryComboBox.setFont(self.genFont)

        self.entryAMPMComboBox = QComboBox()
        entryAMPMComboItems = ["AM", "PM"]
        self.entryAMPMComboBox.addItems(entryAMPMComboItems)
        self.entryAMPMComboBox.setFixedWidth(80)
        self.entryAMPMComboBox.setMinimumHeight(40)
        self.entryAMPMComboBox.setFont(self.genFont)

        self.missIDButton = QPushButton()
        self.missIDButton.setText("RECORD MISSING INFO")
        self.missIDButton.setFixedSize(180, 40)
        self.missIDButton.clicked.connect(self.recordMissInfoMethod)
        self.missIDButton.setFont(self.genFont)

        self.getrecordButton = QPushButton()
        self.getrecordButton.setText("GET ATTENDEES RECORD")
        self.getrecordButton.setFixedSize(180, 40)
        self.getrecordButton.clicked.connect(self.getAttendeesMethod)
        self.getrecordButton.setFont(self.genFont)

        self.spacerItem1 = QSpacerItem(120, 40, QSizePolicy.Policy.Maximum, QSizePolicy.Policy.Expanding)
        self.spacerItem2 = QSpacerItem(40, 40, QSizePolicy.Policy.Maximum, QSizePolicy.Policy.Expanding)
        self.spacerItem3 = QSpacerItem(120, 40, QSizePolicy.Policy.Maximum, QSizePolicy.Policy.Expanding)

        self.slash1 = QLabel()
        self.slash1.setText(" / ")
        self.slash2 = QLabel()
        self.slash2.setText(" / ")

        self.entryGridLayout.addWidget(self.dateLabel, 0, 0, 1, 1)
        self.entryGridLayout.addWidget(self.month, 0, 1, 1, 1)
        self.entryGridLayout.addWidget(self.slash1, 0, 2, 1, 1)
        self.entryGridLayout.addWidget(self.day, 0, 3, 1, 1)
        self.entryGridLayout.addWidget(self.slash2, 0, 4, 1, 1)
        self.entryGridLayout.addWidget(self.year, 0, 5, 1, 1)
        self.entryGridLayout.addItem(self.spacerItem1, 0, 6, 2, 1)
        self.entryGridLayout.addWidget(self.entryField, 0, 7, 1, 1)
        self.entryGridLayout.addWidget(self.entryButton, 0, 9, 1, 1)
        self.entryGridLayout.addItem(self.spacerItem2, 0, 10, 2, 1)
        self.entryGridLayout.addWidget(self.entryComboBox, 0, 13, 1, 1)
        self.entryGridLayout.addWidget(self.entryAMPMComboBox, 0, 14, 1, 1)
        self.entryGridLayout.addItem(self.spacerItem3, 0, 15, 2, 1)
        self.entryGridLayout.addWidget(self.missIDButton, 0, 17, 1, 1)
        self.entryGridLayout.addWidget(self.getrecordButton, 0, 18, 1, 1)

        # ATTENDEES VIEWTABLE FRAME
        self.viewtableFrame = QFrame()
        self.viewTableHboxLayout = QHBoxLayout(self.viewtableFrame)
        self.viewTableHboxLayout.setContentsMargins(0, 0, 0, 0)
        self.viewtableFrame.setContentsMargins(5, 5, 5, 5)

        self.viewTable = QTableView()
        self.viewTable.setFont(self.genFont)
        self.viewTable.setMinimumHeight(545)
        self.viewTable.setEditTriggers(QAbstractItemView.NoEditTriggers)
        self.viewTableItems = ["ID NUMBER", "NAME", "YEAR", "PROGRAM", self.entryComboBox.currentText()]

        self.viewtableModel = QStandardItemModel()
        self.viewtableModel.setHorizontalHeaderLabels(self.viewTableItems)
        self.viewTable.setModel(self.viewtableModel)
        self.viewTable.setAlternatingRowColors(True)
        self.viewTable.verticalHeader().hide()
        self.viewTable.setFont(self.genFont)
        header = self.viewTable.horizontalHeader()
        header.setSectionResizeMode(0, QHeaderView.ResizeToContents)
        header.setSectionResizeMode(1, QHeaderView.Stretch)
        header.setSectionResizeMode(2, QHeaderView.ResizeToContents)
        header.setSectionResizeMode(3, QHeaderView.ResizeToContents)
        header.setSectionResizeMode(4, QHeaderView.Stretch)

        self.attendanceviewTableMethod()

        self.viewTableHboxLayout.addWidget(self.viewTable)

        self.drakeGridLayout.addWidget(self.entryFrame, 0, 0, 1, 5)
        self.drakeGridLayout.addWidget(self.viewtableFrame, 1, 0, 20, 5)

        self.entryButton.clicked.connect(self.entryButtonMethod)
        self.entryComboBox.currentTextChanged.connect(self.attendanceviewTableMethod)
        self.entryAMPMComboBox.currentTextChanged.connect(self.attendanceviewTableMethod)
        self.month.currentTextChanged.connect(self.attendanceviewTableMethod)
        self.month.currentIndexChanged.connect(self.setDayItemsMethod)
        self.year.currentTextChanged.connect(self.attendanceviewTableMethod)
        self.year.currentIndexChanged.connect(self.setDayItemsMethod)
        self.day.currentTextChanged.connect(self.attendanceviewTableMethod)

    @pyqtSlot()
    def entryButtonMethod(self):
        drakedb = sqlite3.connect("SEAITE_Attendance.db")
        dcursor = drakedb.cursor()
        # CHECK ID NUMBER
        id_num = str(self.entryField.text())
        ID_check = dcursor.execute(f"""SELECT ID_NUM FROM STUDENT_LIST WHERE ID_NUM = '{id_num}' """).fetchone()
        print(ID_check)
        combo = self.entryComboBox.currentText().replace("-", "_").strip()
        ampm = self.entryAMPMComboBox.currentText().strip()
        curr_date = self.date
        curr_time = datetime.now()
        if ID_check == None:
            # print("MISSING OR INCOMPLETE ID NUMBER")
            boldfont = QFont()
            boldfont.setBold(True)
            errormsg = "MISSING OR INCOMPLETE ID NUMBER"
            error = QMessageBox()
            error.setWindowIcon(QIcon("resources/failed.png"))
            error.setIcon(QMessageBox.Critical)
            error.setText(errormsg)
            error.setFont(boldfont)
            error.setWindowTitle("ERROR!")
            error.exec_()
        else:
            self.entryField.setText("")
            if combo == "TIME_IN":
                strexe = f"""INSERT OR IGNORE INTO ATTENDANCE (ID_NUM, DATE, AMPM, {combo}) 
                VALUES ('{ID_check[0]}','{curr_date}', '{ampm}',
                '{curr_time.strftime("%I:%M:%S %p")}')"""
                dcursor.execute(strexe)
                drakedb.commit()
                # print("success")
            elif combo == "TIME_OUT":
                time_in_check = dcursor.execute(f"""SELECT TIME_IN FROM 
                ATTENDANCE WHERE ID_NUM = '{ID_check[0]}' AND AMPM = '{ampm}'""").fetchone()
                if time_in_check is None:
                    strexe = f"""INSERT OR IGNORE INTO ATTENDANCE (ID_NUM, DATE, AMPM, {combo}) 
                    VALUES ('{ID_check[0]}','{curr_date}', '{ampm}',
                    '{curr_time.strftime("%I:%M:%S %p")}')"""
                    dcursor.execute(strexe)
                    drakedb.commit()
                    print("success")
                elif time_in_check is not None:
                    strexe = f"""UPDATE ATTENDANCE SET TIME_OUT = '{curr_time.strftime("%I:%M:%S %p")}'
                    WHERE ID_NUM = '{ID_check[0]}' AND DATE = '{curr_date}' AND AMPM = '{ampm}'"""
                    dcursor.execute(strexe)
                    drakedb.commit()

        drakedb.commit()
        self.viewtableModel.setRowCount(0)
        self.attendanceviewTableMethod()

    def attendanceviewTableMethod(self):
        viewTableItems = ["ID NUMBER", "NAME", "YEAR", "PROGRAM", f"{self.entryComboBox.currentText()}"]
        self.viewtableModel.setHorizontalHeaderLabels(viewTableItems)
        self.date = f"{self.month.currentText()}/{self.day.currentText()}/{self.year.currentText()}"
        drakedb = sqlite3.connect("SEAITE_Attendance.db")
        dcursor = drakedb.cursor()
        combo = self.entryComboBox.currentText().replace("-", "_").strip()
        ampm = self.entryAMPMComboBox.currentText().strip()
        curr_date = self.date

        dcursor.execute("DROP TABLE IF EXISTS ATTENDEES_RECORD")

        dcursor.execute(f"""CREATE TABLE IF NOT EXISTS ATTENDEES_RECORD AS SELECT A.INST_ID AS INST_ID, S.ID_NUM AS 
        ID_NUM, S.FULL_NAME AS STUD_NAME, S.YEAR AS YEAR, S.PROGRAM AS PROGRAM, A.DATE AS DATE,A.AMPM AS AMPM, 
        A.TIME_IN AS TIME_IN, A.TIME_OUT AS TIME_OUT FROM STUDENT_LIST S LEFT JOIN ATTENDANCE A ON S.ID_NUM = 
        A.ID_NUM WHERE NOT A.DATE = "N/A" ORDER BY DATE DESC,{combo} DESC""")

        tabledata = dcursor.execute(f"""SELECT ID_NUM, STUD_NAME, YEAR, PROGRAM, {combo} FROM 
                ATTENDEES_RECORD WHERE AMPM = '{ampm}' AND DATE = '{curr_date}'
                ORDER BY {combo} DESC NULLS LAST""").fetchall()
        tabledata = [list(item) for item in tabledata if item[-1] is not None]

        print(tabledata)

        if tabledata != []:
            self.viewtableModel.setRowCount(len(tabledata))
            self.viewtableModel.setColumnCount(len(tabledata[0]))
            for rowid, row in enumerate(tabledata):
                for colid, col in enumerate(row):
                    data = QStandardItem(col)
                    data.setTextAlignment(Qt.AlignCenter)
                    self.viewtableModel.setItem(rowid, colid, data)
        else:
            self.viewtableModel.setRowCount(0)
        drakedb.commit()
        # self.setDayItemsMethod()

    def setDayItemsMethod(self):
        self.day.clear()
        cal = calendar.Calendar()
        month = int(self.month.currentText())
        year = int(self.year.currentText())
        days = [str(day) for day in cal.itermonthdays(year, month) if day != 0]
        days = list(dict.fromkeys(days))
        self.day.addItems(days)

    def recordMissInfoMethod(self):
        drakedb = sqlite3.connect("SEAITE_Attendance.db")
        dcursor = drakedb.cursor()
        program_list = list(dcursor.execute("SELECT PROGRAM FROM STUDENT_LIST").fetchall())
        program_list = [str(item) for items in program_list for item in items]
        program_list = list(dict.fromkeys(program_list))
        program_list.insert(0, "PROGRAM")
        self.recorderWidget = QWidget()
        self.recorderWidget.setFixedSize(580, 400)
        self.recorderWidget.setWindowTitle("Missing INFO Recorder")
        self.recorderWidget.setWindowIcon(QIcon("resources/search.png"))
        self.recorderMainLayout = QVBoxLayout(self.recorderWidget)
        self.recorderMainLayout.setContentsMargins(10, 10, 10, 10)
        self.recorderMainLayout.setSpacing(20)

        self.recorderFrame = QFrame()
        self.recorderFrameLayout = QGridLayout(self.recorderFrame)
        self.recorderFrameLayout.setContentsMargins(10, 10, 10, 10)

        self.idLabel = QLabel()
        self.idLabel.setText("ID NUMBER: ")
        self.idLabel.setMinimumSize(80,40)
        self.idLabel.setFont(self.genFont)

        self.idLineEdit = QLineEdit()
        self.idLineEdit.setPlaceholderText("ENTER ID NUMBER HERE.")
        self.idLineEdit.setMinimumSize(80,40)
        self.idLineEdit.setMaxLength(10)
        self.idLineEdit.setValidator(QtGui.QRegExpValidator(QRegExp(r"^[0-9]*$")))
        self.idLineEdit.setFont(self.genFont)

        self.lnameLabel = QLabel()
        self.lnameLabel.setText("LAST NAME: ")
        self.lnameLabel.setMinimumSize(80,40)
        self.lnameLabel.setFont(self.genFont)

        self.lnameLineEdit = QLineEdit()
        self.lnameLineEdit.setPlaceholderText("ENTER LAST NAME HERE.")
        self.lnameLineEdit.setMinimumSize(80,40)
        self.lnameLineEdit.setValidator(QtGui.QRegExpValidator(QRegExp(r"^[a-zA-Z ]*$")))
        self.lnameLineEdit.setFont(self.genFont)

        self.fnameLabel = QLabel()
        self.fnameLabel.setText("FIRST NAME: ")
        self.fnameLabel.setMinimumSize(80,40)
        self.fnameLabel.setFont(self.genFont)

        self.fnameLineEdit = QLineEdit()
        self.fnameLineEdit.setPlaceholderText("ENTER FIRST NAME HERE.")
        self.fnameLineEdit.setMinimumSize(80,40)
        self.fnameLineEdit.setValidator(QtGui.QRegExpValidator(QRegExp(r"^[a-zA-Z ]*$")))
        self.fnameLineEdit.setFont(self.genFont)

        self.mnameLabel = QLabel()
        self.mnameLabel.setText("MIDDLE NAME: ")
        self.mnameLabel.setMinimumSize(80,40)
        self.mnameLabel.setFont(self.genFont)

        self.mnameLineEdit = QLineEdit()
        self.mnameLineEdit.setPlaceholderText("ENTER MIDDLE NAME HERE.")
        self.mnameLineEdit.setMinimumSize(80,40)
        self.mnameLineEdit.setValidator(QtGui.QRegExpValidator(QRegExp(r"^[a-zA-Z ]*$")))
        self.mnameLineEdit.setFont(self.genFont)

        self.yearLabel = QLabel()
        self.yearLabel.setText("\tYEAR: ")
        self.yearLabel.setMinimumSize(80,40)
        self.yearLabel.setFont(self.genFont)

        self.yearComboBox = QComboBox()
        year = [str(item) for item in range(1, 6)]
        year.insert(0, "YEAR")
        self.yearComboBox.addItems(year)
        self.yearComboBox.setMinimumSize(80,40)
        self.yearComboBox.setFont(self.genFont)

        self.programLabel = QLabel()
        self.programLabel.setText("\tPROGRAM: ")
        self.programLabel.setMinimumSize(80,40)
        self.programLabel.setFont(self.genFont)

        self.programComboBox = QComboBox()
        self.programComboBox.addItems(program_list)
        self.programComboBox.setMinimumSize(80,40)
        self.programComboBox.setFont(self.genFont)

        self.submitButton = QPushButton()
        self.submitButton.setText("SUBMIT")
        self.submitButton.setMinimumHeight(40)
        self.submitButton.setFixedWidth(160)
        self.submitButton.setFont(self.genFont)
        self.submitButton.clicked.connect(self.submitMethod)

        self.clearButton = QPushButton()
        self.clearButton.setText("CLEAR")
        self.clearButton.setMinimumHeight(40)
        self.clearButton.setFixedWidth(160)
        self.clearButton.setFont(self.genFont)
        self.clearButton.clicked.connect(self.clearMethod)

        self.buttonLayout = QHBoxLayout()
        self.buttonLayout.addWidget(self.submitButton)
        self.buttonLayout.addWidget(self.clearButton)

        self.promptLabel = QLabel()

        self.recorderFrameLayout.addWidget(self.idLabel, 0, 0, 1, 5)
        self.recorderFrameLayout.addWidget(self.idLineEdit, 0, 1, 1, 6)
        self.recorderFrameLayout.addWidget(self.lnameLabel, 1, 0, 1, 5)
        self.recorderFrameLayout.addWidget(self.lnameLineEdit, 1, 1, 1, 6)
        self.recorderFrameLayout.addWidget(self.fnameLabel, 2, 0, 1, 5)
        self.recorderFrameLayout.addWidget(self.fnameLineEdit, 2, 1, 1, 6)
        self.recorderFrameLayout.addWidget(self.mnameLabel, 3, 0, 1, 5)
        self.recorderFrameLayout.addWidget(self.mnameLineEdit, 3, 1, 1, 6)
        self.recorderFrameLayout.addWidget(self.yearLabel, 4, 0, 1, 1)
        self.recorderFrameLayout.addWidget(self.yearComboBox, 4, 1, 1, 1)
        self.recorderFrameLayout.addWidget(self.programLabel, 4, 2, 1, 2)
        self.recorderFrameLayout.addWidget(self.programComboBox, 4, 4, 1, 3)
        self.recorderFrameLayout.addWidget(self.promptLabel, 5, 1, 1, 5)
        self.recorderFrameLayout.addLayout(self.buttonLayout, 6, 1, 1, 5)

        self.recorderMainLayout.addWidget(self.recorderFrame)
        self.recorderWidget.show()

    def closeEvent(self, event):
        try:
            if self.recorderWidget:
                self.recorderWidget.close()
        except:
            pass

    def submitMethod(self):
        print("clicked")
        drakedb = sqlite3.connect("SEAITE_Attendance.db")
        dcursor = drakedb.cursor()

        id_num = self.idLineEdit.text()
        print(self.lnameLineEdit.text())
        fullname = f"{self.lnameLineEdit.text().upper()}, {self.fnameLineEdit.text().upper()} {self.mnameLineEdit.text().upper()}"
        year = self.yearComboBox.currentText()
        program = self.programComboBox.currentText()

        combo = self.entryComboBox.currentText().replace("-", "_").strip()
        ampm = self.entryAMPMComboBox.currentText().strip()

        id_check = dcursor.execute(f"""SELECT ID_NUM FROM STUDENT_LIST WHERE ID_NUM = '{id_num}'""").fetchone()
        print(id_check)
        curr_date = self.date
        curr_time = datetime.now()
        curr_time_f = curr_time.strftime("%I:%M:%S %p")

        try:
            if year == "YEAR":
                raise Exception
            elif program == "PROGRAM":
                raise Exception
            elif self.fnameLineEdit.text() == "":
                raise Exception
            elif self.lnameLineEdit.text() == "":
                raise Exception
            elif self.mnameLineEdit.text() == "":
                raise Exception
            elif id_check is not None:
                raise Exception
            else:
                strexe = f"""INSERT OR IGNORE INTO STUDENT_LIST (ID_NUM, FULL_NAME, YEAR, PROGRAM) 
                VALUES ("{id_num}","{fullname}","{year}","{program}")"""
                dcursor.execute(strexe)
                drakedb.commit()

                if combo == "TIME_IN":
                    strexe = f"""INSERT OR IGNORE INTO ATTENDANCE (ID_NUM, DATE, AMPM, {combo}) 
                    VALUES ('{id_num}','{curr_date}', '{ampm}', '{curr_time_f}')"""
                    dcursor.execute(strexe)
                    drakedb.commit()
                    print("success")

                elif combo == "TIME_OUT":
                    strexe = f"""INSERT OR IGNORE INTO STUDENT_LIST (ID_NUM, FULL_NAME, YEAR, PROGRAM) 
                    VALUES ("{id_num}","{fullname}","{year}","{program}")"""
                    dcursor.execute(strexe)
                    drakedb.commit()

                    time_in_check = dcursor.execute(f"""SELECT TIME_IN FROM 
                    ATTENDANCE WHERE ID_NUM = '{id_num}' """).fetchone()
                    if time_in_check != None:
                        strexe = f"""UPDATE ATTENDANCE SET TIME_OUT = '{curr_time_f}'
                        WHERE ID_NUM = '{id_num}' AND DATE = '{curr_date}' AND AMPM = '{ampm}'"""
                        dcursor.execute(strexe)
                        drakedb.commit()

                    else:
                        strexe = f"""INSERT OR IGNORE INTO ATTENDANCE (ID_NUM, DATE, AMPM, {combo}) 
                        VALUES ('{id_num}','{curr_date}', '{ampm}',
                        '{curr_time_f}')"""
                        dcursor.execute(strexe)
                        drakedb.commit()
                        print("success")
                drakedb.commit()

                self.attendanceviewTableMethod()
                self.idLineEdit.clear()
                self.lnameLineEdit.clear()
                self.fnameLineEdit.clear()
                self.mnameLineEdit.clear()
                self.yearComboBox.setCurrentText("YEAR")
                self.programComboBox.setCurrentText("PROGRAM")
                self.promptLabel.setText("SUCCESSFULLY RECORDED!")
        except:
            boldfont = QFont()
            boldfont.setBold(True)
            errormsg = "INVALID INPUT!"
            error = QMessageBox()
            error.setWindowIcon(QIcon("resources/failed.png"))
            error.setIcon(QMessageBox.Critical)
            error.setText(errormsg)
            error.setFont(boldfont)
            error.setWindowTitle("ERROR!")
            error.exec_()

    def clearMethod(self):
        self.attendanceviewTableMethod()
        self.idLineEdit.clear()
        self.lnameLineEdit.clear()
        self.fnameLineEdit.clear()
        self.mnameLineEdit.clear()
        self.yearComboBox.setCurrentText("YEAR")
        self.programComboBox.setCurrentText("PROGRAM")
        self.promptLabel.setText("")

    def getAttendeesMethod(self):
        try:
            curr_date = self.date
            curr_time = datetime.now()
            curr_time_f = curr_time.strftime("%I:%M:%S %p")
            drakedb = sqlite3.connect("SEAITE_Attendance.db")
            dcursor = drakedb.cursor()
            program_list = list(dcursor.execute("SELECT PROGRAM FROM ATTENDEES_RECORD").fetchall())
            program_list = [str(item) for items in program_list for item in items]
            program_list = list(dict.fromkeys(program_list))
            print(program_list)
            if program_list:
                workbook = Workbook()
                desktop = os.path.join(os.path.join(os.environ['USERPROFILE']), 'Desktop')
            elif not program_list:
                raise IndexError

            columnNames = ["ID NUMBER", "FULL NAME", "YEAR", "TIME-IN", "TIME-OUT"]
            for prog in program_list:
                workbook.create_sheet(str(prog))
                sheet = workbook[f'{prog}']
                sheet.column_dimensions['A'].width = 20
                sheet.column_dimensions['B'].width = 50
                sheet.column_dimensions['C'].width = 10
                sheet.column_dimensions['D'].width = 20
                sheet.column_dimensions['E'].width = 20

                scell = sheet.cell(row=1, column=1)
                scell.value = f"DATE: {self.date} AM TIME IN AND OUT"
                scell = sheet.cell(row=2, column=1)
                scell.value = "AM SESSION"

                for id, columns in enumerate(columnNames):
                    scell = sheet.cell(row=3, column=id + 1)
                    scell.value = columns
                AM_attendees = list(dcursor.execute(f"""SELECT ID_NUM, STUD_NAME, YEAR, TIME_IN, TIME_OUT 
                FROM ATTENDEES_RECORD WHERE AMPM = 'AM' AND DATE = '{self.date}' AND PROGRAM = '{prog}' 
                ORDER BY YEAR ASC"""))
                AM_attendees = [list(elem) for elem in AM_attendees]
                for row, rowitem in enumerate(AM_attendees):
                    for col, colitem in enumerate(rowitem):
                        scell = sheet.cell(row=4 + row, column=col + 1)
                        if colitem is None:
                            scell.value = "N/A"
                        else:
                            scell.value = colitem
                print(len(AM_attendees))
                lastrow = len(AM_attendees) + 6
                scell = sheet.cell(row=lastrow, column=1)
                scell.value = "PM SESSION"
                for id, columns in enumerate(columnNames):
                    scell = sheet.cell(row=lastrow + 1, column=id + 1)
                    scell.value = columns

                PM_attendees = list(dcursor.execute(f"""SELECT ID_NUM, STUD_NAME, YEAR, TIME_IN, TIME_OUT 
                FROM ATTENDEES_RECORD WHERE AMPM = 'PM' AND DATE = '{self.date}' AND PROGRAM = '{prog}' 
                ORDER BY YEAR ASC"""))

                PM_attendees = [list(elem) for elem in PM_attendees]
                for row, rowitem in enumerate(PM_attendees):
                    for col, colitem in enumerate(rowitem):
                        scell = sheet.cell(row=lastrow + row + 2, column=col + 1)
                        if colitem is None:
                            scell.value = "N/A"
                        else:
                            scell.value = colitem

            del workbook['Sheet']

            workbook.save(desktop + f"\\seaite_Attendees_Record_{os.environ['USERNAME']}_{curr_date.replace('/','_')}_{curr_time_f.replace(':','_')}.xlsx")
            boldfont = QFont()
            boldfont.setBold(True)
            errormsg = "RECORD SUCCESSFULLY RETRIEVED!"
            error = QMessageBox()
            error.setWindowIcon(QIcon("resources/success.png"))
            error.setIcon(QMessageBox.Information)
            error.setText(errormsg)
            error.setFont(boldfont)
            error.setWindowTitle("SUCCESS!")
            error.exec_()

        except PermissionError:
            boldfont = QFont()
            boldfont.setBold(True)
            errormsg = "PLEASE CLOSE FIRST THE EXCEL FILE TO UPDATE DATA!"
            error = QMessageBox()
            error.setWindowIcon(QIcon("resources/failed.png"))
            error.setIcon(QMessageBox.Critical)
            error.setText(errormsg)
            error.setFont(boldfont)
            error.setWindowTitle("ERROR!")
            error.exec_()

        except IndexError:
            boldfont = QFont()
            boldfont.setBold(True)
            errormsg = "THERE IS NO RECOREDED DATA YET!"
            error = QMessageBox()
            error.setWindowIcon(QIcon("resources/failed.png"))
            error.setIcon(QMessageBox.Critical)
            error.setText(errormsg)
            error.setFont(boldfont)
            error.setWindowTitle("ERROR!")
            error.exec_()
            
if __name__ == "__main__":
    app = QApplication(sys.argv)
    apply_stylesheet(app,theme='light_maroon.xml')
    drake = drakeMain()
    drake.show()
    app.exec_()
