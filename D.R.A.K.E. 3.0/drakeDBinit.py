import sqlite3
import openpyxl
import  os
path = os.getcwd() + "\\SEAITE-Department-Student-List.xlsx"
book = openpyxl.load_workbook(path)
sheet = book.active
maxrow = sheet.max_row

id_num = [sheet.cell(row=r, column=1).value for r in range(2, maxrow + 1)]
stud_name = [sheet.cell(row=r, column=2).value for r in range(2, maxrow + 1)]
year = [sheet.cell(row=r, column=3).value for r in range(2, maxrow + 1)]
program = [sheet.cell(row=r, column=4).value for r in range(2, maxrow + 1)]

def drakeDbfetch():
    # print(len(id_num), len(stud_name), len(year), len(program))

    # print(id_num[0], stud_name[0], year[0], program[0])
    # print(id_num[-1], stud_name[-1], year[-1], program[-1])
    drakedb = sqlite3.connect("SEAITE_Attendance.db")
    dcursor = drakedb.cursor()
    dcursor.execute("""CREATE TABLE IF NOT EXISTS STUDENT_LIST(ID_NUM TEXT PRIMARY KEY, FULL_NAME TEXT, YEAR TEXT, 
    PROGRAM TEXT)""")
    programs = []
    for num, item in enumerate(zip(id_num, stud_name, year, program)):
        strexe = f"""INSERT OR IGNORE INTO STUDENT_LIST (ID_NUM, FULL_NAME, YEAR, PROGRAM) 
        VALUES ("{item[0]}","{item[1]}","{item[2]}","{item[3].split()[0]}") """
        programs.append(item[3].split()[0])
        # print(strexe)
        dcursor.execute(strexe)
    programs = list(dict.fromkeys(programs))
    dcursor.execute("""CREATE TABLE IF NOT EXISTS ATTENDANCE(INST_ID INTEGER PRIMARY KEY AUTOINCREMENT, ID_NUM TEXT NOT 
    NULL, DATE TEXT , AMPM TEXT NOT NULL, TIME_IN TEXT,TIME_OUT TEXT 
    ,  FOREIGN KEY (ID_NUM) REFERENCES STUDENT_LIST (ID_NUM), UNIQUE(ID_NUM, DATE, AMPM)) """)

    drakedb.commit()

if __name__ == "__main__":
    drakeDbfetch()